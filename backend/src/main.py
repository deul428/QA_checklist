from fastapi import FastAPI, Depends, HTTPException, status
from fastapi.middleware.cors import CORSMiddleware
from fastapi.security import OAuth2PasswordBearer, OAuth2PasswordRequestForm
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import and_, func
from datetime import datetime, date, timedelta
from typing import List, Optional
import os
import io
from dotenv import load_dotenv

from services.database import get_db, engine, Base
from models.models import (
    User,
    System,
    CheckItem,
    UserSystemAssignment,
    ChecklistRecord,
    ChecklistRecordLog,
    SubstituteAssignment,
    SubstituteAssignmentLog,
    AdminLog,
)
from services.schemas import (
    UserLogin,
    UserResponse,
    SystemResponse,
    CheckItemResponse,
    ChecklistSubmit,
    ChecklistRecordResponse,
    PasswordChange,
    TestEmailSchedule,
    ConsoleStatsResponse,
    ConsoleFailItemResponse,
    ExcelExportRequest,
    SubstituteAssignmentCreate,
    SubstituteAssignmentResponse,
    CheckItemCreate,
    CheckItemUpdate,
    AssignmentCreate,
)
from services.auth import (
    verify_password,
    get_password_hash,
    create_access_token,
    get_current_user,
)
from services.scheduler import init_scheduler, get_korea_today

load_dotenv()

app = FastAPI(title="DX본부 시스템 체크리스트", version="1.0.0")


# 애플리케이션 시작 시 데이터베이스 테이블 생성 및 스케줄러 초기화
@app.on_event("startup")
async def startup_event():
    try:
        Base.metadata.create_all(bind=engine)
        print("데이터베이스 연결 성공 및 테이블 생성 완료")
    except Exception as e:
        print(f"데이터베이스 연결 오류: {e}")
        print("데이터베이스가 설정되지 않았습니다. .env 파일을 확인하세요.")

    # 스케줄러 초기화 (startup 이벤트에서 실행)
    try:
        init_scheduler()
        print("스케줄러 초기화 완료")
    except Exception as e:
        print(f"스케줄러 초기화 오류: {e}")
        import traceback

        traceback.print_exc()


# CORS 설정
# 개발 환경: 모든 localhost 포트 허용
allowed_origins = [
    "http://localhost:3003",
    "http://localhost:5173",
    "http://127.0.0.1:3003",
    "http://127.0.0.1:5173",
    "http://192.10.10.206:3003",
    "http://192.10.10.206:5173",
]

# 환경 변수에서 추가 origin 허용
if os.getenv("CORS_ORIGINS"):
    additional_origins = [
        origin.strip() for origin in os.getenv("CORS_ORIGINS").split(",")
    ]
    allowed_origins.extend(additional_origins)

# 개발 환경에서는 일반적인 개발 포트들을 모두 허용
# React 개발 서버는 보통 3000-3010, Vite는 5173 등 사용
for port in range(3000, 3011):
    allowed_origins.append(f"http://localhost:{port}")
    allowed_origins.append(f"http://127.0.0.1:{port}")
    allowed_origins.append(f"http://192.10.10.206:{port}")

# Vite 기본 포트
allowed_origins.append("http://localhost:5173")
allowed_origins.append("http://127.0.0.1:5173")
allowed_origins.append("http://192.10.10.206:5173")

# 운영 환경에서는 특정 도메인만 허용하도록 주의
is_production = os.getenv("ENV") == "production"

app.add_middleware(
    CORSMiddleware,
    allow_origins=allowed_origins,
    allow_credentials=True,
    allow_methods=["GET", "POST", "PUT", "DELETE", "OPTIONS", "PATCH"],
    allow_headers=["*"],
    expose_headers=["*"],
)

oauth2_scheme = OAuth2PasswordBearer(tokenUrl="api/auth/login")


@app.get("/")
async def root():
    return {"message": "DX본부 시스템 체크리스트 API"}


@app.get("/api/scheduler/status")
async def get_scheduler_status():
    """스케줄러 상태 확인"""
    from services.scheduler import scheduler

    jobs = []
    if scheduler.running:
        for job in scheduler.get_jobs():
            jobs.append(
                {
                    "id": job.id,
                    "name": job.name,
                    "next_run_time": (
                        job.next_run_time.isoformat() if job.next_run_time else None
                    ),
                    "trigger": str(job.trigger),
                }
            )

    return {"running": scheduler.running, "jobs": jobs}


@app.delete("/api/scheduler/jobs/{job_id}")
async def cancel_scheduled_job(job_id: str):
    """예약된 작업 취소 (개발자용)"""
    from services.scheduler import scheduler

    try:
        if not scheduler.running:
            raise HTTPException(
                status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
                detail="스케줄러가 실행되지 않았습니다.",
            )

        # 작업이 존재하는지 확인
        job = scheduler.get_job(job_id)
        if not job:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail=f"작업 ID '{job_id}'를 찾을 수 없습니다.",
            )

        job_name = job.name
        # 작업 취소
        scheduler.remove_job(job_id)

        return {
            "message": f"작업 '{job_name}' (ID: {job_id})이 취소되었습니다.",
            "job_id": job_id,
        }
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"작업 취소 중 오류 발생: {str(e)}",
        )


@app.post("/api/scheduler/test")
async def test_scheduler():
    """스케줄러 수동 테스트 (관리자용)"""
    from services.scheduler import check_unchecked_items

    try:
        check_unchecked_items()
        return {"message": "스케줄러 테스트가 성공적으로 실행되었습니다."}
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"스케줄러 테스트 중 오류 발생: {str(e)}",
        )


@app.post("/api/scheduler/test-email")
async def test_email_send(schedule: TestEmailSchedule):
    """테스트 메일 스케줄링 (실제 DB의 담당자 이메일 주소 사용)

    시간을 지정하면 해당 시간에 메일을 발송하도록 스케줄링합니다.
    실제 DB의 담당자 이메일 주소로 발송됩니다.
    """
    from services.scheduler import schedule_test_email, scheduler
    import pytz

    try:
        # 시간 유효성 검사
        if not (0 <= schedule.hour <= 23 and 0 <= schedule.minute <= 59):
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="시간은 0-23시, 분은 0-59분 사이여야 합니다.",
            )

        # 스케줄러가 실행 중인지 확인
        if not scheduler.running:
            raise HTTPException(
                status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
                detail="스케줄러가 실행되지 않았습니다.",
            )

        # 스케줄링
        job_id, scheduled_time = schedule_test_email(schedule.hour, schedule.minute)

        kst = pytz.timezone("Asia/Seoul")
        scheduled_time_str = scheduled_time.astimezone(kst).strftime(
            "%Y-%m-%d %H:%M:%S"
        )

        return {
            "message": f"테스트 메일이 {scheduled_time_str}에 발송되도록 스케줄링되었습니다.",
            "scheduled_time": scheduled_time_str,
            "job_id": job_id,
            "note": "실제 DB의 담당자 이메일 주소로 발송됩니다.",
        }
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"테스트 메일 스케줄링 중 오류 발생: {str(e)}",
        )


@app.post("/api/scheduler/test-email-now")
async def test_email_send_now():
    """테스트 메일 즉시 발송 (실제 DB의 담당자 이메일 주소 사용)

    즉시 메일을 발송합니다. 스케줄링 없이 바로 전송됩니다.
    실제 DB의 담당자 이메일 주소로 발송됩니다.
    """
    from services.scheduler import send_test_email_scheduled

    try:
        # 즉시 발송
        send_test_email_scheduled()

        return {
            "message": "메일이 발송되었습니다.",
            "sent_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "note": "실제 DB의 담당자 이메일 주소로 발송되었습니다.",
        }
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"메일 발송 중 오류 발생: {str(e)}",
        )


@app.post("/api/auth/login", response_model=dict)
async def login(
    form_data: OAuth2PasswordRequestForm = Depends(), db: Session = Depends(get_db)
):
    """사번 기반 로그인"""
    user = db.query(User).filter(User.user_id == form_data.username).first()

    # if not user or not verify_password(form_data.password, user.password_hash):
    #     raise HTTPException(
    #         status_code=status.HTTP_401_UNAUTHORIZED,
    #         detail="사번 또는 비밀번호가 올바르지 않습니다",
    #         headers={"WWW-Authenticate": "Bearer"},
    #     )

    access_token = create_access_token(data={"sub": user.user_id})

    # UserResponse 스키마를 사용하여 모든 필드 포함
    from services.schemas import UserResponse

    user_response = UserResponse(
        user_id=user.user_id,
        user_name=user.user_name,
        user_email=user.user_email,
        division=user.division,
        general_headquarters=user.general_headquarters,
        headquarters=user.department,  # department를 headquarters로도 반환 (프론트엔드 호환성)
        department=user.department,
        position=user.position,
        role=user.role,
        console_role=user.console_role,
    )

    return {
        "access_token": access_token,
        "token_type": "bearer",
        "user": user_response.model_dump(),
    }


@app.get("/api/user/me", response_model=UserResponse)
async def get_current_user_info(
    current_user: User = Depends(get_current_user), db: Session = Depends(get_db)
):
    """현재 로그인한 사용자 정보"""
    return current_user


@app.get("/api/user/search", response_model=List[UserResponse])
async def search_users(
    query: Optional[str] = None,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """사용자 검색 (사번, 이름으로 검색)"""
    if not query or len(query.strip()) < 1:
        # 검색어가 없으면 빈 배열 반환
        return []

    search_term = f"%{query.strip()}%"
    users = (
        db.query(User)
        .filter((User.user_id.like(search_term)) | (User.user_name.like(search_term)))
        .limit(50)  # 최대 50명까지만 반환
        .all()
    )

    return users


@app.post("/api/user/change-password", response_model=dict)
async def change_password(
    password_data: PasswordChange,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """비밀번호 변경"""
    # 현재 비밀번호 확인
    if not verify_password(password_data.current_password, current_user.password_hash):
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="현재 비밀번호가 올바르지 않습니다",
        )

    # 새 비밀번호로 변경
    current_user.password_hash = get_password_hash(password_data.new_password)
    db.commit()

    return {"message": "비밀번호가 성공적으로 변경되었습니다"}


def validate_environment_for_system(
    system: System, environment: str
) -> None:
    """시스템의 환경 존재 여부 검증
    
    Args:
        system: 시스템 객체
        environment: 검증할 환경 ('dev', 'stg', 'prd')
    
    Raises:
        HTTPException: 시스템이 해당 환경을 지원하지 않는 경우
    """
    if environment == "dev" and not system.has_dev:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"시스템 '{system.system_name}'은(는) dev 환경을 지원하지 않습니다.",
        )
    if environment == "stg" and not system.has_stg:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"시스템 '{system.system_name}'은(는) stg 환경을 지원하지 않습니다.",
        )
    if environment == "prd" and not system.has_prd:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"시스템 '{system.system_name}'은(는) prd 환경을 지원하지 않습니다.",
        )


def check_system_access(
    user_id: str, system_id: int, db: Session, check_date: Optional[date] = None
) -> bool:
    """시스템 접근 권한 확인 (일반 담당자 또는 활성화된 대체 담당자)

    Args:
        user_id: 확인할 사용자 user_id
        system_id: 시스템 ID
        db: 데이터베이스 세션
        check_date: 확인할 날짜 (None이면 오늘 날짜)

    Returns:
        접근 권한이 있으면 True, 없으면 False
    """
    if check_date is None:
        check_date = date.today()

    # 1. 일반 담당자 권한 확인
    assignment = (
        db.query(UserSystemAssignment)
        .filter(
            UserSystemAssignment.user_id == user_id,
            UserSystemAssignment.system_id == system_id,
        )
        .first()
    )

    if assignment:
        return True

    # 2. 대체 담당자 권한 확인 (기간 내에 활성화되어 있는지)
    substitute = (
        db.query(SubstituteAssignment)
        .filter(
            SubstituteAssignment.substitute_user_id == user_id,
            SubstituteAssignment.system_id == system_id,
            SubstituteAssignment.start_date <= check_date,
            SubstituteAssignment.end_date >= check_date,
        )
        .first()
    )

    return substitute is not None


@app.get("/api/user/systems", response_model=List[SystemResponse])
async def get_user_systems(
    current_user: User = Depends(get_current_user), db: Session = Depends(get_db)
):
    """사용자가 담당하는 시스템 목록 (일반 담당자 + 활성화된 대체 담당자)"""
    today = date.today()

    # 일반 담당자 시스템
    assignments = (
        db.query(UserSystemAssignment)
        .filter(UserSystemAssignment.user_id == current_user.user_id)
        .all()
    )
    system_ids = {assignment.system_id for assignment in assignments}

    # 활성화된 대체 담당자 시스템
    substitutes = (
        db.query(SubstituteAssignment)
        .filter(
            SubstituteAssignment.substitute_user_id == current_user.user_id,
            SubstituteAssignment.start_date <= today,
            SubstituteAssignment.end_date >= today,
        )
        .all()
    )
    system_ids.update({substitute.system_id for substitute in substitutes})

    systems = db.query(System).filter(System.system_id.in_(system_ids)).all()

    return systems


@app.get("/api/systems/{system_id}/check-items", response_model=List[CheckItemResponse])
async def get_check_items(
    system_id: int,
    environment: str = "prd",  # 'dev', 'stg', 'prd'
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """시스템의 체크 항목 목록 (특이사항 포함)"""
    # 환경 유효성 검사
    if environment not in ["dev", "stg", "prd"]:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="환경은 'dev', 'stg', 'prd' 중 하나여야 합니다.",
        )
    
    # 권한 확인 (일반 담당자 또는 활성화된 대체 담당자)
    if not check_system_access(current_user.user_id, system_id, db):
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="해당 시스템에 대한 접근 권한이 없습니다",
        )

    check_items = (
        db.query(CheckItem)
        .filter(
            CheckItem.system_id == system_id,
            CheckItem.status == "active",  # 삭제된 항목 제외
        )
        .order_by(CheckItem.item_id)
        .all()
    )
    
    # 시스템이 해당 환경을 지원하는지 확인
    system = db.query(System).filter(System.system_id == system_id).first()
    if system:
        # 시스템이 지원하지 않는 환경이면 빈 리스트 반환
        if environment == "dev" and not system.has_dev:
            check_items = []
        elif environment == "stg" and not system.has_stg:
            check_items = []
        elif environment == "prd" and not system.has_prd:
            check_items = []

    return check_items


@app.get("/api/checklist/today", response_model=List[ChecklistRecordResponse])
async def get_today_checklist(
    environment: str = "prd",  # 'dev', 'stg', 'prd'
    current_user: User = Depends(get_current_user), db: Session = Depends(get_db)
):
    """오늘 날짜의 체크리스트 기록 조회

    확인자가 여러 명인 경우, 한 명이 체크하면 다른 확인자들도 체크된 것으로 보임.
    따라서 user_id 필터 없이 check_item_id와 check_date만으로 조회.
    """
    # 환경 유효성 검사
    if environment not in ["dev", "stg", "prd"]:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="환경은 'dev', 'stg', 'prd' 중 하나여야 합니다.",
        )
    
    today = date.today()

    # 사용자가 담당하는 시스템의 체크 항목 ID 목록 (일반 담당자 + 활성화된 대체 담당자)
    # 일반 담당자 시스템
    assignments = (
        db.query(UserSystemAssignment)
        .filter(
            UserSystemAssignment.user_id == current_user.user_id,
            UserSystemAssignment.environment == environment,
        )
        .all()
    )
    system_ids = {a.system_id for a in assignments}

    # 활성화된 대체 담당자 시스템
    substitutes = (
        db.query(SubstituteAssignment)
        .filter(
            SubstituteAssignment.substitute_user_id == current_user.user_id,
            SubstituteAssignment.start_date <= today,
            SubstituteAssignment.end_date >= today,
        )
        .all()
    )
    system_ids.update({substitute.system_id for substitute in substitutes})

    check_item_ids = [
        item.item_id
        for item in db.query(CheckItem)
        .filter(
            CheckItem.system_id.in_(system_ids),
            CheckItem.status == "active",  # 삭제된 항목 제외
        )
        .all()
    ]

    # 오늘 날짜에 체크된 기록 조회 (다른 사람이 체크한 것도 포함)
    records = (
        db.query(ChecklistRecord)
        .filter(
            ChecklistRecord.check_item_id.in_(check_item_ids),
            ChecklistRecord.check_date == today,
            ChecklistRecord.environment == environment,
        )
        .all()
    )

    return records


@app.post("/api/checklist/submit", response_model=dict)
async def submit_checklist(
    checklist_data: ChecklistSubmit,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """체크리스트 제출 (PASS/FAIL 저장)"""
    today = date.today()

    for item in checklist_data.items:
        # 체크 항목 권한 확인
        check_item = (
            db.query(CheckItem).filter(CheckItem.item_id == item.check_item_id).first()
        )
        if not check_item:
            continue

        # 권한 확인 (일반 담당자 또는 활성화된 대체 담당자)
        if not check_system_access(
            current_user.user_id, check_item.system_id, db, today
        ):
            continue

        # 환경 유효성 검사
        item_environment = getattr(item, "environment", "prd")
        if item_environment not in ["dev", "stg", "prd"]:
            continue
        
        # 시스템의 환경 존재 여부 검증 (데이터 무결성 보장)
        system = db.query(System).filter(System.system_id == check_item.system_id).first()
        if system:
            try:
                validate_environment_for_system(system, item_environment)
            except HTTPException:
                # 시스템이 해당 환경을 지원하지 않으면 스킵
                continue

        # 기존 기록 확인 및 업데이트 또는 생성
        # 확인자가 여러 명인 경우, 한 명이 체크하면 다른 사람도 체크된 것으로 보임.
        # 따라서 user_id 필터 없이 check_item_id, check_date, environment로 확인.
        existing_record = (
            db.query(ChecklistRecord)
            .filter(
                ChecklistRecord.check_item_id == item.check_item_id,
                ChecklistRecord.check_date == today,
                ChecklistRecord.environment == item_environment,
            )
            .first()
        )

        if existing_record:
            # 기존 기록이 있으면 업데이트 (누가 체크했는지는 기록)
            old_status = existing_record.status
            existing_record.status = item.status
            existing_record.fail_notes = item.fail_notes
            existing_record.checked_at = datetime.now()
            existing_record.system_id = check_item.system_id  # system_id 업데이트 (일관성 유지)
            # 체크한 사람 정보도 업데이트 (같은 사람이 다시 체크한 경우)
            existing_record.user_id = current_user.user_id

            # 로그 기록 (UPDATE 액션)
            log_entry = ChecklistRecordLog(
                user_id=current_user.user_id,
                check_item_id=item.check_item_id,
                system_id=check_item.system_id,  # 시스템 ID 추가
                check_date=today,
                environment=item_environment,
                status=item.status,
                fail_notes=item.fail_notes,
                action="UPDATE",
            )
            db.add(log_entry)
        else:
            # 새 기록 생성
            new_record = ChecklistRecord(
                user_id=current_user.user_id,
                check_item_id=item.check_item_id,
                system_id=check_item.system_id,  # 시스템 ID 추가
                check_date=today,
                environment=item_environment,
                status=item.status,
                fail_notes=item.fail_notes,
            )
            db.add(new_record)

            # 로그 기록 (CREATE 액션)
            log_entry = ChecklistRecordLog(
                user_id=current_user.user_id,
                check_item_id=item.check_item_id,
                system_id=check_item.system_id,  # 시스템 ID 추가
                check_date=today,
                environment=item_environment,
                status=item.status,
                fail_notes=item.fail_notes,
                action="CREATE",
            )
            db.add(log_entry)

    db.commit()
    return {"message": "체크리스트가 성공적으로 저장되었습니다"}


def check_console_access(current_user: User):
    """console 페이지 접근 권한 체크 (DB의 console_role 컬럼 확인)"""
    if not current_user.console_role:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="console 페이지 접근 권한이 없습니다",
        )


@app.get("/api/console/stats", response_model=ConsoleStatsResponse)
async def get_console_stats(
    environment: Optional[str] = None,  # 'dev', 'stg', 'prd' 또는 None (모든 환경)
    current_user: User = Depends(get_current_user), db: Session = Depends(get_db)
):
    """console 페이지 통계 조회 (오늘 날짜 기준 pass/fail/미점검)"""
    check_console_access(current_user)

    today = get_korea_today()

    # 오늘 체크된 항목
    record_query = db.query(ChecklistRecord).filter(ChecklistRecord.check_date == today)
    if environment:
        record_query = record_query.filter(ChecklistRecord.environment == environment)
    checked_records = record_query.all()

    pass_count = sum(1 for r in checked_records if r.status == "PASS")
    fail_count = sum(1 for r in checked_records if r.status == "FAIL")

    # 모든 활성 체크 항목 수 (삭제된 항목 제외)
    all_items_count = db.query(CheckItem).filter(CheckItem.status == "active").count()
    unchecked_count = all_items_count - len(checked_records)

    return ConsoleStatsResponse(
        pass_count=pass_count,
        fail_count=fail_count,
        unchecked_count=unchecked_count,
    )


@app.get("/api/console/fail-items", response_model=List[ConsoleFailItemResponse])
async def get_console_fail_items(
    environment: Optional[str] = None,  # 'dev', 'stg', 'prd' 또는 None (모든 환경)
    current_user: User = Depends(get_current_user), db: Session = Depends(get_db)
):
    """console 페이지 fail 항목 목록 조회

    오늘 날짜에 FAIL 상태가 된 모든 항목을 조회합니다.
    - 처음부터 FAIL인 경우
    - PASS에서 FAIL로 변경된 경우
    - FAIL에서 PASS로 변경되었다가 다시 FAIL로 변경된 경우
    """
    check_console_access(current_user)

    today = get_korea_today()

    # 오늘 날짜의 모든 로그를 시간순으로 조회
    log_query = (
        db.query(ChecklistRecordLog)
        .filter(ChecklistRecordLog.check_date == today)
    )
    if environment:
        log_query = log_query.filter(ChecklistRecordLog.environment == environment)
    all_logs = log_query.order_by(ChecklistRecordLog.created_at).all()  # 전체 로그를 시간순으로 정렬

    # 각 check_item_id별로 상태 변경 추적 (시간순으로 정렬된 로그를 항목별로 그룹화)
    item_status_history = (
        {}
    )  # {check_item_id: [(status, created_at, user_id, fail_notes), ...]}

    for log in all_logs:
        if log.check_item_id not in item_status_history:
            item_status_history[log.check_item_id] = []
        item_status_history[log.check_item_id].append(
            (log.status, log.created_at, log.user_id, log.fail_notes)
        )

    # 각 항목의 로그를 시간순으로 정렬 (혹시 모를 정렬 문제 방지)
    for check_item_id in item_status_history:
        item_status_history[check_item_id].sort(
            key=lambda x: x[1]
        )  # created_at 기준 정렬

    result = []

    for check_item_id, history in item_status_history.items():
        # 체크 항목 정보
        check_item = db.query(CheckItem).filter(CheckItem.item_id == check_item_id).first()
        if not check_item:
            continue

        # 시스템 정보
        system = db.query(System).filter(System.system_id == check_item.system_id).first()
        if not system:
            continue

        # 최종 상태 확인 (시간순으로 정렬된 마지막 로그의 상태)
        # history는 이미 시간순으로 정렬되어 있음
        final_status = history[-1][0]  # 마지막 로그의 상태

        # 최종 상태가 FAIL인 항목만 표시
        if final_status != "FAIL":
            continue

        # 첫 번째 FAIL 로그 찾기 (PASS -> FAIL 또는 처음부터 FAIL)
        first_fail_log = None
        for status, created_at, user_id, fail_notes in history:
            if status == "FAIL":
                first_fail_log = (status, created_at, user_id, fail_notes)
                break

        if not first_fail_log:
            continue

        # 사용자 정보 (첫 번째 FAIL을 기록한 사용자)
        user = db.query(User).filter(User.user_id == first_fail_log[2]).first()
        if not user:
            continue

        # FAIL 이후에 PASS로 변경되었는지 확인
        first_fail_time = first_fail_log[1]
        pass_after_fail = None
        for status, created_at, user_id, fail_notes in history:
            if status == "PASS" and created_at > first_fail_time:
                pass_after_fail = (created_at, fail_notes)
                break

        # PASS 이후에 다시 FAIL로 변경되었는지 확인
        is_resolved = False
        resolved_date = None
        resolved_time = None

        if pass_after_fail:
            # PASS 이후에 다시 FAIL이 있는지 확인
            pass_time = pass_after_fail[0]
            fail_after_pass = None
            for status, created_at, user_id, fail_notes in history:
                if status == "FAIL" and created_at > pass_time:
                    fail_after_pass = (created_at, fail_notes)
                    break

            if not fail_after_pass:
                # PASS 이후에 FAIL이 없으면 해결된 것으로 간주
                # 하지만 최종 상태가 FAIL이므로, 이 경우는 발생하지 않아야 함
                # 최종 상태가 FAIL이면 미해결로 처리
                is_resolved = False
                resolved_date = None
                resolved_time = None

        # 최신 FAIL 로그 정보 (최종 FAIL 상태)
        latest_fail_log = None
        for status, created_at, user_id, fail_notes in reversed(history):
            if status == "FAIL":
                latest_fail_log = (created_at, fail_notes)
                break

        if not latest_fail_log:
            latest_fail_log = (first_fail_log[1], first_fail_log[3])

        # 환경 정보는 로그에서 가져오기 (첫 번째 FAIL 로그의 environment 사용)
        # 로그가 여러 환경에 걸쳐 있을 수 있으므로, 첫 번째 FAIL 로그의 environment 사용
        log_environment = "prd"  # 기본값
        for log in db.query(ChecklistRecordLog).filter(
            ChecklistRecordLog.check_item_id == check_item_id,
            ChecklistRecordLog.status == "FAIL",
            ChecklistRecordLog.check_date == today
        ).order_by(ChecklistRecordLog.created_at).limit(1).all():
            log_environment = log.environment
            break
        
        result.append(
            ConsoleFailItemResponse(
                id=check_item_id,  # check_item_id를 id로 사용
                system_id=system.system_id,
                system_name=system.system_name,
                check_item_id=check_item_id,
                item_name=check_item.item_name,
                environment=log_environment,
                fail_notes=latest_fail_log[1] if latest_fail_log else first_fail_log[3],
                fail_time=first_fail_log[1],  # 첫 번째 FAIL 시간
                user_id=user.user_id,
                user_name=user.user_name,
                is_resolved=is_resolved,
                resolved_date=resolved_date,
                resolved_time=resolved_time,
            )
        )

    # fail_time 기준으로 정렬 (최신순)
    result.sort(key=lambda x: x.fail_time, reverse=True)

    print(f"[Console API] 오늘 FAIL 항목: {len(result)}개")
    for item in result:
        print(
            f"  - {item.item_name}: is_resolved={item.is_resolved}, fail_time={item.fail_time}"
        )

    return result


@app.get("/api/checklist/unchecked", response_model=List[dict])
async def get_unchecked_items(
    environment: Optional[str] = None,  # 'dev', 'stg', 'prd' 또는 None (모든 환경)
    current_user: User = Depends(get_current_user), db: Session = Depends(get_db)
):
    """오늘 체크되지 않은 항목 조회

    확인자가 여러 명인 경우, 한 명이 체크하면 다른 사람도 체크된 것으로 보임.
    따라서 user_id 필터 없이 check_item_id, check_date, environment로 확인.
    """
    today = date.today()

    # 사용자가 담당하는 시스템의 모든 체크 항목 (일반 담당자 + 활성화된 대체 담당자)
    # 일반 담당자 시스템
    assignment_query = db.query(UserSystemAssignment).filter(
        UserSystemAssignment.user_id == current_user.user_id
    )
    if environment:
        assignment_query = assignment_query.filter(
            UserSystemAssignment.environment == environment
        )
    assignments = assignment_query.all()
    system_ids = {a.system_id for a in assignments}

    # 활성화된 대체 담당자 시스템
    substitutes = (
        db.query(SubstituteAssignment)
        .filter(
            SubstituteAssignment.substitute_user_id == current_user.user_id,
            SubstituteAssignment.start_date <= today,
            SubstituteAssignment.end_date >= today,
        )
        .all()
    )
    system_ids.update({substitute.system_id for substitute in substitutes})

    item_query = db.query(CheckItem).filter(
        CheckItem.system_id.in_(system_ids),
        CheckItem.status == "active",  # 삭제된 항목 제외
    )
    if environment:
        item_query = item_query.filter(CheckItem.environment == environment)
    all_items = item_query.all()

    # 오늘 체크된 항목 (다른 사람이 체크한 것도 포함)
    record_query = db.query(ChecklistRecord).filter(ChecklistRecord.check_date == today)
    if environment:
        record_query = record_query.filter(ChecklistRecord.environment == environment)
    checked_records = record_query.all()
    checked_item_ids = {(r.check_item_id, r.environment) for r in checked_records}

    # 체크되지 않은 항목
    unchecked_items = [
        {
            "check_item_id": item.item_id,
            "item_name": item.item_name,
            "system_id": item.system_id,
            "environment": environment,  # 요청된 환경 사용
            "system_name": db.query(System)
            .filter(System.system_id == item.system_id)
            .first()
            .system_name,
        }
        for item in all_items
        if (item.item_id, environment) not in checked_item_ids
    ]

    return unchecked_items


@app.post("/api/console/export-excel")
async def export_excel(
    request: ExcelExportRequest,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """체크리스트 통계 엑셀 다운로드"""
    check_console_access(current_user)

    # openpyxl import 확인
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        from openpyxl.chart import BarChart, Reference
    except ImportError as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"엑셀 생성 라이브러리(openpyxl)가 설치되지 않았습니다. pip install openpyxl (오류: {str(e)})",
        )

    try:

        # 날짜 범위 검증
        if request.start_date > request.end_date:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="시작 날짜가 종료 날짜보다 늦을 수 없습니다.",
            )

        # 날짜 범위 내의 모든 체크리스트 기록 조회
        records = (
            db.query(ChecklistRecord)
            .filter(
                and_(
                    ChecklistRecord.check_date >= request.start_date,
                    ChecklistRecord.check_date <= request.end_date,
                )
            )
            .all()
        )

        print(f"[엑셀 다운로드] 날짜 범위: {request.start_date} ~ {request.end_date}")
        print(f"[엑셀 다운로드] 조회된 기록 수: {len(records)}")

        # 모든 활성 체크 항목 조회 (담당자 정보 포함, 삭제된 항목 제외)
        all_items = db.query(CheckItem).filter(CheckItem.status == "active").all()
        all_systems = db.query(System).all()
        all_assignments = db.query(UserSystemAssignment).all()
        all_users = db.query(User).all()

        print(f"[엑셀 다운로드] 전체 체크 항목 수: {len(all_items)}")
        print(f"[엑셀 다운로드] 전체 시스템 수: {len(all_systems)}")
        print(f"[엑셀 다운로드] 전체 담당자 할당 수: {len(all_assignments)}")

        # 시스템, 항목, 담당자 매핑 생성
        system_map = {s.system_id: s.system_name for s in all_systems}
        item_map = {item.item_id: item for item in all_items}

        # (system_id, item_id, environment) 조합으로 담당자 매핑 생성
        assignment_map = {}
        for assignment in all_assignments:
            key = (assignment.system_id, assignment.item_id, assignment.environment)
            if key not in assignment_map:
                assignment_map[key] = set()  # 중복 제거를 위해 set 사용
            user = next((u for u in all_users if u.user_id == assignment.user_id), None)
            if user:
                assignment_map[key].add(user.user_name)  # 사번 제거, 이름만 사용

        # set을 리스트로 변환 (정렬하여 일관성 유지)
        for key in assignment_map:
            assignment_map[key] = sorted(list(assignment_map[key]))

        user_map = {u.user_id: u.user_name for u in all_users}

        # 엑셀 워크북 생성
        wb = Workbook()
        ws = wb.active
        ws.title = "체크리스트 통계"

        # 스타일 정의
        header_fill = PatternFill(
            start_color="366092", end_color="366092", fill_type="solid"
        )
        header_font = Font(bold=True, color="FFFFFF", size=11)
        border_side = Side(style="thin", color="000000")
        border = Border(
            left=border_side, right=border_side, top=border_side, bottom=border_side
        )
        center_alignment = Alignment(horizontal="center", vertical="center")

        # 헤더 작성
        headers = ["날짜", "시스템", "환경", "항목", "담당자", "상태", "비고"]
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_alignment
            cell.border = border

        # 데이터 작성 (중복 제거: 같은 check_item_id, check_date, environment에 대해 가장 최근 것만 사용)
        # 먼저 (check_item_id, check_date, environment)별로 그룹화하고, 가장 최근 checked_at을 가진 레코드만 선택
        records_dict = {}
        for record in records:
            key = (record.check_item_id, record.check_date, record.environment)
            if key not in records_dict:
                records_dict[key] = record
            else:
                # 같은 키가 있으면 checked_at이 더 최근인 것으로 교체
                if record.checked_at > records_dict[key].checked_at:
                    records_dict[key] = record

        # 실제로 기록이 있는 날짜만 추출
        dates_with_records = set()
        for record in records:
            dates_with_records.add(record.check_date)

        # 날짜 범위 내의 모든 날짜 생성 (실제 기록이 있는 날짜만)
        date_list = sorted(list(dates_with_records))

        # 실제 기록이 있는 날짜에 대해서만 데이터 생성
        # 각 항목에 대해 모든 환경(dev, stg, prd)에 대해 기록 확인
        excel_data = []
        environments = ["dev", "stg", "prd"]
        for check_item in all_items:
            for check_date in date_list:
                for env in environments:
                    key = (check_item.item_id, check_date, env)
                    if key in records_dict:
                        # 체크된 기록이 있는 경우
                        record = records_dict[key]
                        excel_data.append(
                            {
                                "date": check_date,
                                "system_id": check_item.system_id,
                                "item_id": check_item.item_id,
                                "item_name": check_item.item_name,
                                "environment": env,
                                "status": record.status,
                                "fail_notes": record.fail_notes or "",
                            }
                        )
                    # 미점검 항목은 제외 (모든 환경에 대해 생성하면 너무 많아짐)

        # 날짜, 시스템, 환경, 항목 순으로 정렬
        excel_data.sort(key=lambda x: (x["date"], x["system_id"], x["environment"], x["item_name"]))

        row_idx = 2
        for data in excel_data:
            system_name = system_map.get(data["system_id"], "")
            # (system_id, item_id, environment) 조합으로 담당자 조회
            assignment_key = (data["system_id"], data["item_id"], data["environment"])
            responsible_users = ", ".join(assignment_map.get(assignment_key, []))

            ws.cell(
                row=row_idx, column=1, value=data["date"].strftime("%Y-%m-%d")
            ).border = border
            ws.cell(row=row_idx, column=2, value=system_name).border = border
            ws.cell(row=row_idx, column=3, value=data["environment"].upper()).border = border
            ws.cell(row=row_idx, column=4, value=data["item_name"]).border = border
            ws.cell(row=row_idx, column=5, value=responsible_users).border = border
            ws.cell(row=row_idx, column=6, value=data["status"]).border = border
            ws.cell(row=row_idx, column=7, value=data["fail_notes"]).border = border

            row_idx += 1

        print(f"[엑셀 다운로드] 엑셀에 작성된 행 수: {row_idx - 2}")

        # 열 너비 조정
        ws.column_dimensions["A"].width = 12
        ws.column_dimensions["B"].width = 25
        ws.column_dimensions["C"].width = 10
        ws.column_dimensions["D"].width = 40
        ws.column_dimensions["E"].width = 30
        ws.column_dimensions["F"].width = 10
        ws.column_dimensions["G"].width = 50

        # 필터 적용
        ws.auto_filter.ref = f"A1:G{row_idx - 1}"

        # 통계 시트 생성
        stats_ws = wb.create_sheet("통계")

        # 날짜별 통계 계산
        date_stats = {}
        for record in records:
            date_str = record.check_date.strftime("%Y-%m-%d")
            if date_str not in date_stats:
                date_stats[date_str] = {"PASS": 0, "FAIL": 0, "UNCHECKED": 0}

            if record.status == "PASS":
                date_stats[date_str]["PASS"] += 1
            elif record.status == "FAIL":
                date_stats[date_str]["FAIL"] += 1

        # 전체 항목 수 계산
        total_items = len(all_items)

        # 통계 헤더
        stats_headers = ["날짜", "PASS", "FAIL", "미점검", "전체"]
        for col_idx, header in enumerate(stats_headers, 1):
            cell = stats_ws.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_alignment
            cell.border = border

        # 통계 데이터
        stats_row = 2
        sorted_dates = sorted(date_stats.keys())
        for date_str in sorted_dates:
            stats = date_stats[date_str]
            unchecked = total_items - stats["PASS"] - stats["FAIL"]

            stats_ws.cell(row=stats_row, column=1, value=date_str).border = border
            stats_ws.cell(row=stats_row, column=2, value=stats["PASS"]).border = border
            stats_ws.cell(row=stats_row, column=3, value=stats["FAIL"]).border = border
            stats_ws.cell(row=stats_row, column=4, value=unchecked).border = border
            stats_ws.cell(row=stats_row, column=5, value=total_items).border = border

            stats_row += 1

        # 통계 열 너비 조정
        stats_ws.column_dimensions["A"].width = 12
        stats_ws.column_dimensions["B"].width = 10
        stats_ws.column_dimensions["C"].width = 10
        stats_ws.column_dimensions["D"].width = 10
        stats_ws.column_dimensions["E"].width = 10

        # 차트 생성
        if len(sorted_dates) > 0:
            chart = BarChart()
            chart.type = "col"
            chart.style = 2
            chart.title = f"체크리스트 통계 ({request.start_date} ~ {request.end_date})"
            chart.y_axis.title = "개수"
            chart.x_axis.title = "날짜"

            data = Reference(
                stats_ws, min_col=2, min_row=1, max_col=4, max_row=stats_row - 1
            )
            cats = Reference(stats_ws, min_col=1, min_row=2, max_row=stats_row - 1)
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(cats)

            stats_ws.add_chart(chart, "G2")

        # 메모리 버퍼에 엑셀 파일 저장
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        # 파일명 생성 (한글 인코딩 처리)
        filename = f"체크리스트_통계_{request.start_date}_{request.end_date}.xlsx"
        # RFC 5987 형식으로 한글 파일명 인코딩
        from urllib.parse import quote

        encoded_filename = quote(filename.encode("utf-8"))

        return StreamingResponse(
            io.BytesIO(output.read()),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f"attachment; filename*=UTF-8''{encoded_filename}",
            },
        )
    except Exception as e:
        # ImportError는 이미 함수 시작 부분에서 처리됨
        import traceback

        error_detail = traceback.format_exc()
        print(f"[엑셀 다운로드] 오류 발생: {str(e)}")
        print(f"[엑셀 다운로드] 상세 오류:\n{error_detail}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"엑셀 파일 생성 중 오류 발생: {str(e)}",
        )


# ==================== 대체 담당자 관리 API ====================


@app.post("/api/substitute/create", response_model=SubstituteAssignmentResponse)
async def create_substitute_assignment(
    data: SubstituteAssignmentCreate,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """대체 담당자 지정 (시스템 단위)"""
    # 원래 담당자인지 확인
    assignment = (
        db.query(UserSystemAssignment)
        .filter(
            UserSystemAssignment.user_id == current_user.user_id,
            UserSystemAssignment.system_id == data.system_id,
        )
        .first()
    )

    if not assignment:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="해당 시스템의 담당자가 아닙니다.",
        )

    # 날짜 검증
    if data.end_date < data.start_date:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="종료일은 시작일보다 이후여야 합니다.",
        )

    # 대체 담당자 존재 확인
    substitute_user = (
        db.query(User).filter(User.user_id == data.substitute_user_id).first()
    )
    if not substitute_user:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="대체 담당자를 찾을 수 없습니다.",
        )

    # 시스템 존재 확인
    system = db.query(System).filter(System.system_id == data.system_id).first()
    if not system:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="시스템을 찾을 수 없습니다.",
        )

    # 중복 확인 (같은 원래 담당자, 대체 담당자, 시스템, 기간이 겹치는지)
    overlapping = (
        db.query(SubstituteAssignment)
        .filter(
            SubstituteAssignment.original_user_id == current_user.user_id,
            SubstituteAssignment.substitute_user_id == data.substitute_user_id,
            SubstituteAssignment.system_id == data.system_id,
            # 기간이 겹치는지 확인
            SubstituteAssignment.start_date <= data.end_date,
            SubstituteAssignment.end_date >= data.start_date,
        )
        .first()
    )

    if overlapping:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="해당 기간에 이미 대체 담당자가 지정되어 있습니다.",
        )

    # 대체 담당자 생성
    substitute = SubstituteAssignment(
        original_user_id=current_user.user_id,
        substitute_user_id=data.substitute_user_id,
        system_id=data.system_id,
        start_date=data.start_date,
        end_date=data.end_date,
    )
    db.add(substitute)
    db.commit()
    db.refresh(substitute)

    # 대체 담당자 할당 로그 기록
    log_substitute_assignment_change(
        db=db,
        substitute_assignment_id=substitute.id,
        action="CREATE",
        changed_by_user_id=current_user.user_id,
        new_data={
            "original_user_id": substitute.original_user_id,
            "substitute_user_id": substitute.substitute_user_id,
            "system_id": substitute.system_id,
            "start_date": substitute.start_date.isoformat() if substitute.start_date else None,
            "end_date": substitute.end_date.isoformat() if substitute.end_date else None,
        },
    )

    # 대체 담당자에게 메일 발송
    try:
        from services.scheduler import send_email

        # 해당 시스템의 모든 체크 항목 조회
        check_items = (
            db.query(CheckItem)
            .filter(
                CheckItem.system_id == data.system_id,
                CheckItem.status == "active",  # 삭제된 항목 제외
            )
            .order_by(CheckItem.item_id)
            .all()
        )

        # 메일 본문 생성
        items_table_rows = ""
        if check_items:
            # 첫 번째 항목 (시스템명 포함)
            items_table_rows += f"""
                        <tr>
                            <td style="padding: 8px; border: 1px solid #ddd; font-size: 14px; font-weight: bold; vertical-align: top;" rowspan="{len(check_items)}">{system.system_name}</td>
                            <td style="padding: 8px; border: 1px solid #ddd; font-size: 14px;">{check_items[0].item_name}</td>
                            <td style="padding: 8px; border: 1px solid #ddd; font-size: 14px; text-align: center;">{data.start_date.strftime('%Y-%m-%d')}</td>
                            <td style="padding: 8px; border: 1px solid #ddd; font-size: 14px; text-align: center;">{data.end_date.strftime('%Y-%m-%d')}</td>
                        </tr>
            """

            # 나머지 항목 추가
            for item in check_items[1:]:
                items_table_rows += f"""
                        <tr>
                            <td style="padding: 8px; border: 1px solid #ddd; font-size: 14px;">{item.item_name}</td>
                            <td style="padding: 8px; border: 1px solid #ddd; font-size: 14px; text-align: center;">{data.start_date.strftime('%Y-%m-%d')}</td>
                            <td style="padding: 8px; border: 1px solid #ddd; font-size: 14px; text-align: center;">{data.end_date.strftime('%Y-%m-%d')}</td>
                        </tr>
                """
        else:
            # 체크 항목이 없는 경우
            items_table_rows = f"""
                        <tr>
                            <td style="padding: 8px; border: 1px solid #ddd; font-size: 14px; font-weight: bold;">{system.system_name}</td>
                            <td style="padding: 8px; border: 1px solid #ddd; font-size: 14px;">(항목 없음)</td>
                            <td style="padding: 8px; border: 1px solid #ddd; font-size: 14px; text-align: center;">{data.start_date.strftime('%Y-%m-%d')}</td>
                            <td style="padding: 8px; border: 1px solid #ddd; font-size: 14px; text-align: center;">{data.end_date.strftime('%Y-%m-%d')}</td>
                        </tr>
            """

        email_body = f"""
        <html>
            <body>
                <h3>대체 담당자 요청 알림</h3>
                <p>안녕하세요. DX본부 체크리스트 시스템입니다. {substitute_user.user_name} 님께서 다음과 같이 대체 담당자로 지정되었습니다.</p>
                <br />
                <h3>대체 담당 요청자 정보: {current_user.user_name}({current_user.user_id})</h3>
                <p style="font-size: 12px;">귀하를 대체 담당자로 지정한 임직원의 정보입니다.</p>
                <br />
                <h3>대체 담당자 요청 내역</h3>
                <table style="width: 100%; border-collapse: collapse; margin-top: 10px; margin-bottom: 20px;">
                    <thead>
                        <tr style="background-color: #f5f5f5;">
                            <th style="padding: 10px; border: 1px solid #ddd; font-size: 14px; font-weight: bold; text-align: left;">시스템</th>
                            <th style="padding: 10px; border: 1px solid #ddd; font-size: 14px; font-weight: bold; text-align: left;">항목</th>
                            <th style="padding: 10px; border: 1px solid #ddd; font-size: 14px; font-weight: bold; text-align: center;">시작일</th>
                            <th style="padding: 10px; border: 1px solid #ddd; font-size: 14px; font-weight: bold; text-align: center;">종료일</th>
                        </tr>
                    </thead>
                    <tbody>
                        {items_table_rows}
                    </tbody>
                </table>
                <br />
                <p>지정된 기간 동안 해당 시스템의 체크리스트를 작성할 수 있습니다.</p>
                <p>바쁘신 와중에 협조해 주셔서 감사합니다.</p>
            </body>
        </html>
        """

        subject = f"[대체 담당자 요청] {system.system_name} 시스템 체크리스트 대체 담당자 요청 안내"

        # 대체 담당자에게 메일 발송
        if substitute_user.user_email:
            send_email(
                to_email=substitute_user.user_email,
                subject=subject,
                body=email_body,
                cc_emails=[],
            )
            print(
                f"[대체 담당자 메일] {substitute_user.user_name} ({substitute_user.user_email})에게 메일 발송 완료"
            )
        else:
            print(
                f"[대체 담당자 메일] {substitute_user.user_name}의 이메일 주소가 없어 메일을 발송할 수 없습니다."
            )
    except Exception as e:
        # 메일 발송 실패해도 대체 담당자 지정은 성공한 것으로 처리
        print(f"[대체 담당자 메일] 메일 발송 중 오류 발생: {e}")
        import traceback

        traceback.print_exc()

    # 응답 생성
    original_user = db.query(User).filter(User.user_id == current_user.user_id).first()
    return SubstituteAssignmentResponse(
        id=substitute.id,
        original_user_id=substitute.original_user_id,
        original_user_name=original_user.user_name,
        substitute_user_id=substitute.substitute_user_id,
        substitute_user_name=substitute_user.user_name,
        system_id=substitute.system_id,
        system_name=system.system_name,
        start_date=substitute.start_date,
        end_date=substitute.end_date,
        created_at=substitute.created_at,
        is_active=substitute.start_date <= date.today() <= substitute.end_date,
    )


@app.get("/api/substitute/list", response_model=List[SubstituteAssignmentResponse])
async def list_substitute_assignments(
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """내가 지정한 대체 담당자 목록 조회"""
    substitutes = (
        db.query(SubstituteAssignment)
        .filter(SubstituteAssignment.original_user_id == current_user.user_id)
        .order_by(SubstituteAssignment.start_date.desc())
        .all()
    )

    result = []
    today = date.today()
    for sub in substitutes:
        original_user = (
            db.query(User).filter(User.user_id == sub.original_user_id).first()
        )
        substitute_user = (
            db.query(User).filter(User.user_id == sub.substitute_user_id).first()
        )
        system = db.query(System).filter(System.system_id == sub.system_id).first()

        result.append(
            SubstituteAssignmentResponse(
                id=sub.id,
                original_user_id=sub.original_user_id,
                original_user_name=original_user.user_name if original_user else "",
                substitute_user_id=sub.substitute_user_id,
                substitute_user_name=substitute_user.user_name if substitute_user else "",
                system_id=sub.system_id,
                system_name=system.system_name if system else "",
                start_date=sub.start_date,
                end_date=sub.end_date,
                created_at=sub.created_at,
                is_active=sub.start_date <= today <= sub.end_date,
            )
        )

    return result


@app.get("/api/substitute/active", response_model=List[SubstituteAssignmentResponse])
async def get_active_substitute_assignments(
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """내가 대체 담당자로 지정된 활성화된 목록 조회"""
    today = date.today()
    substitutes = (
        db.query(SubstituteAssignment)
        .filter(
            SubstituteAssignment.substitute_user_id == current_user.user_id,
            SubstituteAssignment.start_date <= today,
            SubstituteAssignment.end_date >= today,
        )
        .order_by(SubstituteAssignment.start_date.desc())
        .all()
    )

    result = []
    for sub in substitutes:
        original_user = (
            db.query(User).filter(User.user_id == sub.original_user_id).first()
        )
        substitute_user = (
            db.query(User).filter(User.user_id == sub.substitute_user_id).first()
        )
        system = db.query(System).filter(System.system_id == sub.system_id).first()

        result.append(
            SubstituteAssignmentResponse(
                id=sub.id,
                original_user_id=sub.original_user_id,
                original_user_name=original_user.user_name if original_user else "",
                substitute_user_id=sub.substitute_user_id,
                substitute_user_name=substitute_user.user_name if substitute_user else "",
                system_id=sub.system_id,
                system_name=system.system_name if system else "",
                start_date=sub.start_date,
                end_date=sub.end_date,
                created_at=sub.created_at,
                is_active=True,
            )
        )

    return result


@app.delete("/api/substitute/{substitute_id}")
async def delete_substitute_assignment(
    substitute_id: int,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """대체 담당자 삭제 (원래 담당자만 삭제 가능)"""
    substitute = (
        db.query(SubstituteAssignment)
        .filter(SubstituteAssignment.id == substitute_id)
        .first()
    )

    if not substitute:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="대체 담당자를 찾을 수 없습니다.",
        )

    # 원래 담당자인지 확인
    if substitute.original_user_id != current_user.user_id:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="대체 담당자를 삭제할 권한이 없습니다.",
        )

    # 변경 전 데이터 저장
    old_data = {
        "original_user_id": substitute.original_user_id,
        "substitute_user_id": substitute.substitute_user_id,
        "system_id": substitute.system_id,
        "start_date": substitute.start_date.isoformat() if substitute.start_date else None,
        "end_date": substitute.end_date.isoformat() if substitute.end_date else None,
    }

    # 대체 담당자 할당 로그 기록 (삭제 전에 기록)
    log_substitute_assignment_change(
        db=db,
        substitute_assignment_id=substitute.id,
        action="DELETE",
        changed_by_user_id=current_user.user_id,
        old_data=old_data,
    )

    db.delete(substitute)
    db.commit()

    return {"message": "대체 담당자가 삭제되었습니다."}


# ==================== 관리자 관리 API ====================


def log_substitute_assignment_change(
    db: Session,
    substitute_assignment_id: Optional[int],
    action: str,
    changed_by_user_id: str,
    old_data: Optional[dict] = None,
    new_data: Optional[dict] = None,
):
    """대체 담당자 할당 변경 로그 기록"""
    import json

    substitute_log = SubstituteAssignmentLog(
        substitute_assignment_id=substitute_assignment_id,
        action=action,
        changed_by_user_id=changed_by_user_id,
        old_data=json.dumps(old_data, ensure_ascii=False) if old_data else None,
        new_data=json.dumps(new_data, ensure_ascii=False) if new_data else None,
    )
    db.add(substitute_log)
    db.commit()


def log_admin_action(
    db: Session,
    admin_user_id: int,
    action: str,
    entity_type: str,
    entity_id: Optional[int],
    old_data: Optional[dict] = None,
    new_data: Optional[dict] = None,
):
    """관리자 작업 로그 기록"""
    import json

    admin_log = AdminLog(
        admin_user_id=admin_user_id,
        action=action,
        entity_type=entity_type,
        entity_id=entity_id,
        old_data=json.dumps(old_data, ensure_ascii=False) if old_data else None,
        new_data=json.dumps(new_data, ensure_ascii=False) if new_data else None,
    )
    db.add(admin_log)
    db.commit()


@app.get("/api/admin/systems", response_model=List[SystemResponse])
async def get_all_systems(
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """모든 시스템 목록 조회 (관리자용)"""
    check_console_access(current_user)

    systems = db.query(System).order_by(System.system_name).all()
    return systems


@app.get("/api/admin/check-items", response_model=List[CheckItemResponse])
async def get_admin_check_items(
    system_id: Optional[int] = None,
    environment: Optional[str] = None,  # 'dev', 'stg', 'prd' 또는 None (모든 환경)
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """체크 항목 목록 조회 (관리자용, 삭제된 항목 포함)"""
    check_console_access(current_user)

    query = db.query(CheckItem)
    if system_id:
        query = query.filter(CheckItem.system_id == system_id)

    items = query.order_by(CheckItem.system_id, CheckItem.item_id).all()
    
    # environment 파라미터는 무시 (더 이상 사용하지 않음)
    return items


@app.post("/api/admin/check-items", response_model=CheckItemResponse)
async def create_check_item(
    data: CheckItemCreate,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """체크 항목 추가 (관리자용)"""
    check_console_access(current_user)

    # 시스템 존재 확인
    system = db.query(System).filter(System.system_id == data.system_id).first()
    if not system:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="시스템을 찾을 수 없습니다.",
        )

    # 중복 항목 확인 (같은 system_id, item_name 조합)
    existing_item = db.query(CheckItem).filter(
        CheckItem.system_id == data.system_id,
        CheckItem.item_name == data.item_name
    ).first()
    
    if existing_item:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"이미 같은 이름의 항목이 존재합니다: '{data.item_name}'",
        )
    
    # 항목 생성 (environment 없이)
    check_item = CheckItem(
        system_id=data.system_id,
        item_name=data.item_name,
        item_description=data.item_description,
        status="active",
    )
    db.add(check_item)
    db.commit()
    db.refresh(check_item)

    # 관리자 작업 로그 기록
    log_admin_action(
        db=db,
        admin_user_id=current_user.user_id,
        action="CREATE",
        entity_type="check_item",
        entity_id=check_item.item_id,
        new_data={
            "system_id": check_item.system_id,
            "item_name": check_item.item_name,
            "item_description": check_item.item_description,
        },
    )

    return check_item


@app.put("/api/admin/check-items/{item_id}", response_model=CheckItemResponse)
async def update_check_item(
    item_id: int,
    data: CheckItemUpdate,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """체크 항목 수정 (관리자용)"""
    check_console_access(current_user)

    check_item = db.query(CheckItem).filter(CheckItem.item_id == item_id).first()
    if not check_item:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="체크 항목을 찾을 수 없습니다.",
        )

    # 변경 전 데이터 저장
    old_data = {
        "system_id": check_item.system_id,
        "item_name": check_item.item_name,
        "item_description": check_item.item_description,
        "status": check_item.status,
    }

    # 데이터 업데이트
    if data.item_name is not None:
        check_item.item_name = data.item_name
    # item_description은 None이 아니면 업데이트 (빈 문자열도 명시적으로 업데이트)
    # Pydantic에서 Optional 필드에 빈 문자열을 보내면 빈 문자열로 처리됨
    if data.item_description is not None:
        check_item.item_description = data.item_description
    if data.status is not None:
        if data.status not in ["active", "deleted"]:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="status는 'active' 또는 'deleted'만 가능합니다.",
            )
        check_item.status = data.status

    db.commit()
    db.refresh(check_item)

    # 변경 후 데이터 저장
    new_data = {
        "system_id": check_item.system_id,
        "item_name": check_item.item_name,
        "item_description": check_item.item_description,
        "status": check_item.status,
    }

    # 관리자 작업 로그 기록
    log_admin_action(
        db=db,
        admin_user_id=current_user.user_id,
        action="UPDATE",
        entity_type="check_item",
        entity_id=check_item.item_id,
        old_data=old_data,
        new_data=new_data,
    )

    return check_item


@app.delete("/api/admin/check-items/{item_id}")
async def delete_check_item(
    item_id: int,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """체크 항목 삭제 (관리자용, soft delete)"""
    check_console_access(current_user)

    check_item = db.query(CheckItem).filter(CheckItem.item_id == item_id).first()
    if not check_item:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="체크 항목을 찾을 수 없습니다.",
        )

    if check_item.status == "deleted":
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="이미 삭제된 항목입니다.",
        )

    # 변경 전 데이터 저장
    old_data = {
        "system_id": check_item.system_id,
        "item_name": check_item.item_name,
        "item_description": check_item.item_description,
        "status": check_item.status,
    }

    # Soft delete (status를 'deleted'로 변경)
    check_item.status = "deleted"
    db.commit()

    # 변경 후 데이터 저장
    new_data = {
        "system_id": check_item.system_id,
        "item_name": check_item.item_name,
        "item_description": check_item.item_description,
        "status": "deleted",
    }

    # 관리자 작업 로그 기록
    log_admin_action(
        db=db,
        admin_user_id=current_user.user_id,
        action="DELETE",
        entity_type="check_item",
        entity_id=check_item.item_id,
        old_data=old_data,
        new_data=new_data,
    )

    return {"message": "체크 항목이 삭제되었습니다."}


@app.get("/api/admin/users", response_model=List[UserResponse])
async def get_all_users(
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """모든 사용자 목록 조회 (관리자용, 담당자 배정용)"""
    check_console_access(current_user)

    users = db.query(User).order_by(User.user_name).all()
    return users


@app.post("/api/admin/assignments")
async def create_assignments(
    data: AssignmentCreate,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """담당자 배정 (관리자용, 여러 명 가능)"""
    check_console_access(current_user)

    # 체크 항목 확인
    check_item = db.query(CheckItem).filter(CheckItem.item_id == data.check_item_id).first()
    if not check_item:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="체크 항목을 찾을 수 없습니다.",
        )

    if check_item.status == "deleted":
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="삭제된 항목에는 담당자를 배정할 수 없습니다.",
        )

    # 시스템 확인
    if check_item.system_id != data.system_id:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="시스템 ID가 일치하지 않습니다.",
        )
    
    # 환경 유효성 검사
    assignment_environment = getattr(data, "environment", "prd")
    if assignment_environment not in ["dev", "stg", "prd"]:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="환경은 'dev', 'stg', 'prd' 중 하나여야 합니다.",
        )
    
    # 시스템의 환경 존재 여부 검증 (데이터 무결성 보장)
    system = db.query(System).filter(System.system_id == check_item.system_id).first()
    if system:
        validate_environment_for_system(system, assignment_environment)

    created_assignments = []
    for user_id in data.user_ids:
        # 사용자 확인
        user = db.query(User).filter(User.user_id == user_id).first()
        if not user:
            continue

        # 중복 확인
        existing = (
            db.query(UserSystemAssignment)
            .filter(
                UserSystemAssignment.user_id == user_id,
                UserSystemAssignment.system_id == data.system_id,
                UserSystemAssignment.item_id == data.check_item_id,
                UserSystemAssignment.environment == assignment_environment,
            )
            .first()
        )

        if existing:
            continue

        # 배정 생성
        assignment = UserSystemAssignment(
            user_id=user_id,
            system_id=data.system_id,
            item_id=data.check_item_id,
            environment=assignment_environment,
        )
        db.add(assignment)
        created_assignments.append(assignment)

    db.commit()

    # 로그 기록
    for assignment in created_assignments:
        log_admin_action(
            db=db,
            admin_user_id=current_user.user_id,
            action="CREATE",
            entity_type="assignment",
            entity_id=assignment.id,
            new_data={
                "user_id": assignment.user_id,
                "system_id": assignment.system_id,
                "item_id": assignment.item_id,
                "environment": assignment.environment,
            },
        )

    return {
        "message": f"{len(created_assignments)}명의 담당자가 배정되었습니다.",
        "created_count": len(created_assignments),
    }


@app.delete("/api/admin/assignments/{assignment_id}")
async def delete_assignment(
    assignment_id: int,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """담당자 배정 삭제 (관리자용)"""
    check_console_access(current_user)

    assignment = (
        db.query(UserSystemAssignment)
        .filter(UserSystemAssignment.id == assignment_id)
        .first()
    )

    if not assignment:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="담당자 배정을 찾을 수 없습니다.",
        )

    # 변경 전 데이터 저장
    old_data = {
        "user_id": assignment.user_id,
        "system_id": assignment.system_id,
        "item_id": assignment.item_id,
    }

    db.delete(assignment)
    db.commit()

    # 로그 기록
    log_admin_action(
        db=db,
        admin_user_id=current_user.user_id,
        action="DELETE",
        entity_type="assignment",
        entity_id=assignment_id,
        old_data=old_data,
    )

    return {"message": "담당자 배정이 삭제되었습니다."}


@app.get("/api/admin/assignments")
async def get_assignments(
    system_id: Optional[int] = None,
    check_item_id: Optional[int] = None,
    environment: Optional[str] = None,  # 'dev', 'stg', 'prd' 또는 None (모든 환경)
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """담당자 배정 목록 조회 (관리자용)"""
    check_console_access(current_user)

    query = db.query(UserSystemAssignment)
    if system_id:
        query = query.filter(UserSystemAssignment.system_id == system_id)
    if check_item_id:
        query = query.filter(UserSystemAssignment.item_id == check_item_id)
    if environment:
        query = query.filter(UserSystemAssignment.environment == environment)

    assignments = query.order_by(
        UserSystemAssignment.system_id, UserSystemAssignment.environment, UserSystemAssignment.item_id
    ).all()

    result = []
    for assignment in assignments:
        user = db.query(User).filter(User.user_id == assignment.user_id).first()
        system = db.query(System).filter(System.system_id == assignment.system_id).first()
        check_item = db.query(CheckItem).filter(CheckItem.item_id == assignment.item_id).first()
        result.append(
            {
                "id": assignment.id,
                "user_id": assignment.user_id,
                "user_name": user.user_name if user else "",
                "system_id": assignment.system_id,
                "system_name": system.system_name if system else "",
                "item_id": assignment.item_id,
                "item_name": check_item.item_name if check_item else "",
                "environment": assignment.environment,
                "created_at": assignment.created_at,
            }
        )

    return result


if __name__ == "__main__":
    import uvicorn

    uvicorn.run(app, host="0.0.0.0", port=8003)
